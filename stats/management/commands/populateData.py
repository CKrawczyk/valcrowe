from django.core.management.base import BaseCommand
from stats.models import *
from openpyxl import load_workbook
from ._private import *

wb_eth = load_workbook(
    filename='./inital_data/ORIGINAL DATA IDS ETHNICITY.xlsx',
    data_only=True
)
ws_eth = wb_eth['All']

wb = load_workbook(
    filename='./inital_data/VOLCROWE DATA.xlsx',
    data_only=True
)
ws = wb['All_VOLCROWE DATA']


class Command(BaseCommand):
    def populateStatic(self):
        QTList = ['OP', 'QU', 'YN', 'AD', 'GN', 'ET', 'ED']
        ADList = ['PE', 'Va', 'Ca', 'Un', 'RC', 'So', 'Co', 'WL', 'SC', 'Lo', 'Qu', 'In', 'Em', 'Ed', 'Ti', 'FF', 'Re']
        self.QT = {}
        self.AD = {}
        self.questions = {}

        for i in QTList:
            qt = QuestionType.objects.filter(kind__exact=i)
            if len(qt):
                qt = qt[0]
            else:
                qt = QuestionType(kind=i)
                qt.save()
            self.QT[i] = qt

        for j in ADList:
            ad = ADQuestionCategory.objects.filter(category__exact=j)
            if len(ad):
                ad = ad[0]
            else:
                ad = ADQuestionCategory(category=j)
                ad.save()
            self.AD[j] = ad

        for qdx, q in enumerate(questionsList):
            number = qdx + 1
            question = Question.objects.filter(number__exact=number)
            if len(question):
                question = question[0]
            else:
                kwargs = {
                    'number': number,
                    'question_text': q['text'],
                    'kind': self.QT[q['kind']],
                    'plot_type': q['plot_type'],
                }
                if q['context'] is not None:
                    context = ADQuestionContext.objects.filter(
                        context__exact=q['context']
                    )
                    if len(context):
                        kwargs['context'] = context[0]
                    else:
                        qc = ADQuestionContext(
                            context=q['context']
                        )
                        qc.save()
                        kwargs['context'] = qc
                if q['category'] is not None:
                    kwargs['category'] = self.AD[q['category']]
                question = Question(**kwargs)
                question.save()
            self.questions[number] = question

    def createUser(self, idx, values):
        info = {'id': idx}
        info['talk_posts'] = values[3]
        info['duration_first_last_talk_days'] = values[4]
        info['total_n_classifications'] = values[5]
        info['total_n_sessions'] = values[6]
        info['total_n_projects'] = values[7]
        info['total_unique_days'] = values[8]
        info['first_classification'] = values[9]
        info['last_classification'] = values[10]
        info['duration_first_last_talk_hours'] = values[11]
        info['duration_first_last_classification_hours'] = values[12]
        info['min_classifications_per_session'] = values[13]
        info['max_classifications_per_session'] = values[14]
        info['median_classifications_per_session'] = values[15]
        info['mean_classifications_per_session'] = values[16]
        info['mean_duration_classification_hours'] = values[17]
        info['median_duration_classification_hours'] = values[18]
        info['mean_duration_session_hours'] = values[19]
        info['median_duration_session_hours'] = values[20]
        info['min_duration_session_hours'] = values[21]
        info['max_duration_session_hours'] = values[22]
        info['mean_duration_session_first2_hours'] = values[23]
        info['mean_duration_session_last2_hours'] = values[24]
        info['mean_duration_classification_first2_hours'] = values[25]
        info['mean_duration_classification_last2_hours'] = values[26]
        info['min_number_projects_per_session'] = values[27]
        info['max_number_projects_per_session'] = values[28]
        info['median_number_projects_per_session'] = values[29]
        info['mean_number_projects_per_session'] = values[30]
        if values[49]:
            info['country'] = 'US'
        elif values[50]:
            info['country'] = 'UK'
        elif values[51]:
            info['country'] = 'Ca'
        elif values[52]:
            info['country'] = 'Au'
        elif values[53]:
            info['country'] = 'Ge'
        elif values[54]:
            info['country'] = 'Fr'
        elif values[55]:
            info['country'] = 'Ne'
        elif values[56]:
            info['country'] = 'Po'
        elif values[57]:
            info['country'] = 'In'
        u = User(**info)
        u.save()
        return u

    def createProjects(self, user, values):
        for pdx, p in enumerate(values[1]):
            project_info = {
                'project': projectLookup[p],
                'classifications': values[2][pdx],
                'home_project': p in values[0],
                'user': user,
            }
            projects = Projects(**project_info)
            projects.save()

    def createSurveyProject(self, user, values):
        survey_info = {'user': user}
        if values[32]:
            survey_info['project'] = 'GZ'
        elif values[33]:
            survey_info['project'] = 'PH'
        elif values[34]:
            survey_info['project'] = 'PW'
        elif values[35]:
            survey_info['project'] = 'SE'
        elif values[36]:
            survey_info['project'] = 'SS'
        survey_info['total_n_classifications'] = values[37]
        survey_info['project_duration_1_days'] = values[38]
        survey_info['project_duration_2_days'] = values[39]
        survey_info['total_n_sessions'] = values[40]
        survey_info['max_classifications_per_session'] = values[41]
        survey_info['mean_classifications_per_session'] = values[42]
        survey_info['project_duration_hours'] = values[43]
        survey_info['total_n_unique_days'] = values[44]
        survey_info['mean_duration_session_hours'] = values[45]
        survey_info['longest_active_session_hours'] = values[46]
        survey_info['longest_inactive_session_hours'] = values[47]
        survey_info['mean_duration_classification_hours'] = values[48]
        surveyProject = SurveyProject(**survey_info)
        surveyProject.save()

    def createAnswer(self, user, question, answer_info):
        kind = question.kind.kind
        answer = Answer(question=question, user=user)
        answer.save()
        answer_info['answerID'] = answer
        if kind == 'OP':
            answerValue = AnswerOpen(**answer_info)
        elif kind == 'QU':
            answerValue = AnswerQuiz(**answer_info)
        elif kind == 'YN':
            answerValue = AnswerBool(**answer_info)
        elif kind == 'AD':
            answerValue = AnswerAgree(**answer_info)
        elif kind == 'GN':
            answerValue = AnswerGender(**answer_info)
        elif kind == 'ET':
            answerValue = AnswerEthnicity(**answer_info)
        elif kind == 'ED':
            answerValue = AnswerEducationLevel(**answer_info)
        answerValue.save()

    def handle(self, *args, **options):
        self.populateStatic()
        for rdx, r in enumerate(ws.rows):
            if (rdx > 2) and (rdx < 1918) and (rdx != 146) and (rdx != 401):
                values = [i.value for i in r]
                for i in range(len(values)):
                    if (type(values[i]) == unicode) and ('[' in values[i]):
                        try:
                            exec 'values[i] = ' + values[i]
                        except:
                            pass
                    if type(values[i]) == long:
                        values[i] = int(values[i])
                index = rdx - 2
                user = User.objects.filter(id__exact=index)
                if len(user):
                    user = user[0]
                else:
                    user = self.createUser(index, values)
                    self.createProjects(user, values)
                    self.createSurveyProject(user, values)
                    for qdx in range(1, 70):
                        question = self.questions[qdx]
                        ans = {'answer': values[qdx + 57]}
                        self.createAnswer(user, question, ans)
                    # science quiz
                    ans = {
                        'score': values[127],
                        'percent': values[128],
                        'maxScore': 15,
                        'confidence': values[129],
                    }
                    self.createAnswer(user, self.questions[70], ans)
                    # project quiz
                    ans = {
                        'score': values[135],
                        'percent': values[136],
                        'maxScore': 15,
                        'confidence': values[137],
                    }
                    self.createAnswer(user, self.questions[71], ans)
                    # ploitics quiz
                    ans = {
                        'score': values[138],
                        'percent': values[139],
                        'maxScore': 15,
                        'confidence': values[140],
                    }
                    self.createAnswer(user, self.questions[72], ans)
                    # Gender
                    ans = {
                        'answer': 'M' if values[141] else 'F'
                    }
                    self.createAnswer(user, self.questions[73], ans)
                    # age
                    ans = {'answer': values[142]}
                    self.createAnswer(user, self.questions[74], ans)
                    # Ethnicity
                    ethnicity = ws_eth['EY{0}'.format(rdx)].value
                    specify = ws_eth['EZ{0}'.format(rdx)].value
                    ans = {
                        'answer': ethnicityLookup[ethnicity],
                    }
                    if specify != '':
                        ans['specify'] = specify
                    self.createAnswer(user, self.questions[75], ans)
                    for qdx in range(76, 100):
                        question = self.questions[qdx]
                        ans = {'answer': values[qdx + 68]}
                        self.createAnswer(user, question, ans)
